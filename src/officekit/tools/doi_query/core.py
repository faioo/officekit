"""Business logic for querying DOI values from an Excel file."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import requests

CROSSREF_WORKS_URL = "https://api.crossref.org/works"
REQUIRED_COLUMNS = ("Title", "Journal", "Year")


@dataclass(frozen=True)
class DOIQuerySummary:
    """Summary of a DOI enrichment run."""

    output_path: Path
    total: int
    success: int
    not_found: int
    errors: int


def query_doi(
    title: Any,
    journal: Any,
    year: Any,
    *,
    timeout: int = 30,
) -> str:
    """Query Crossref for the most likely DOI for a paper."""
    query = " ".join(str(value).strip() for value in (title, journal, year) if value)
    if not query:
        return "Not Found"

    try:
        response = requests.get(
            CROSSREF_WORKS_URL,
            params={"query.bibliographic": query, "rows": 1},
            timeout=timeout,
        )
        response.raise_for_status()
    except requests.exceptions.Timeout:
        return "Timeout"
    except requests.exceptions.RequestException as error:
        return f"Error: {str(error)[:80]}"

    items = response.json().get("message", {}).get("items", [])
    if not items:
        return "Not Found"

    return items[0].get("DOI") or "Not Found"


def enrich_excel_with_doi(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    sheet_name: str | None = None,
    timeout: int = 30,
    progress_callback: Callable[[int, int, str, str], None] | None = None,
) -> DOIQuerySummary:
    """Read an Excel file, append a DOI column, and save a new workbook."""
    from typing import Callable
    source = Path(input_path).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"Excel file not found: {source}")

    destination = (
        Path(output_path).expanduser().resolve()
        if output_path
        else source.with_name(f"{source.stem}_with_doi{source.suffix}")
    )
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(source)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in header_cells]
        columns = _find_columns(headers)

        doi_col = _find_column(headers, "DOI") or len(headers) + 1
        sheet.cell(row=1, column=doi_col, value="DOI")

        total = max(sheet.max_row - 1, 0)
        success = 0
        errors = 0

        for row_index in range(2, sheet.max_row + 1):
            title = sheet.cell(row=row_index, column=columns["Title"]).value
            journal = sheet.cell(row=row_index, column=columns["Journal"]).value
            year = sheet.cell(row=row_index, column=columns["Year"]).value

            doi = query_doi(title, journal, year, timeout=timeout)
            sheet.cell(row=row_index, column=doi_col, value=doi)

            if doi.startswith("Error") or doi == "Timeout":
                errors += 1
            elif doi != "Not Found":
                success += 1

            if progress_callback:
                progress_callback(row_index - 1, total, str(title or ""), doi)

        workbook.save(destination)
    finally:
        workbook.close()

    return DOIQuerySummary(
        output_path=destination,
        total=total,
        success=success,
        not_found=total - success - errors,
        errors=errors,
    )


def _find_columns(headers: list[Any]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_name in REQUIRED_COLUMNS:
        column_index = _find_column(headers, column_name)
        if column_index is not None:
            columns[column_name] = column_index

    missing = [name for name in REQUIRED_COLUMNS if name not in columns]
    if missing:
        raise ValueError(
            "Excel file is missing required columns: "
            f"{', '.join(missing)}. Current headers: {headers}"
        )

    return columns


def _find_column(headers: list[Any], name: str) -> int | None:
    for index, header in enumerate(headers, start=1):
        if header and str(header).strip().lower() == name.lower():
            return index

    return None
